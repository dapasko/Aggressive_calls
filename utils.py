import os
import uuid
import pandas as pd
import io
import logging
from openpyxl.utils import get_column_letter
from config import TEMP_DIR


def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Приводит DataFrame к безопасному виду для записи в Excel."""
    # Сначала приводим строковые колонки к чистому string-type
    for col in ('description', 'education_program'):
        if col in df.columns:
            df[col] = (
                df[col]
                .replace(0, '')    # убрать возможные цифровые 0
                .fillna('')        # заменить NaN пустыми строками
                .astype(str)       # явно строковый тип
            )

    # Обработка остальных колонок
    for col in df.columns:
        if col in ('description', 'education_program'):
            continue

        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%d.%m.%Y %H:%M:%S').fillna('')
        elif pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(0)
        else:
            df[col] = df[col].astype(str).replace('nan', '').fillna('')
    return df

def save_temp_file(df: pd.DataFrame) -> str:
    """Сохраняет DataFrame во временный файл и возвращает его ID."""
    temp_id = str(uuid.uuid4())
    temp_path = os.path.join(TEMP_DIR, f"{temp_id}.xlsx")

    try:
        # Очищаем DataFrame перед записью
        df = sanitize_dataframe(df.copy())

        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='assign')
            ws = writer.sheets['assign']
            for idx, col in enumerate(df.columns, start=1):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                ws.column_dimensions[get_column_letter(idx)].width = max_len

        # Скрываем файл (только Windows)
        if os.name == 'nt':
            import ctypes
            ctypes.windll.kernel32.SetFileAttributesW(temp_path, 0x02)

        logging.info(f"Временный файл сохранен: {temp_path}")
        return temp_id
    except Exception as e:
        logging.error(f"Ошибка при сохранении временного файла: {e}")
        raise


def load_temp_file(temp_id: str) -> pd.DataFrame:
    """Загружает DataFrame из временного файла по ID."""
    temp_path = os.path.join(TEMP_DIR, f"{temp_id}.xlsx")
    if not os.path.exists(temp_path):
        raise FileNotFoundError(f"Файл {temp_id} не найден")
    return pd.read_excel(temp_path)


def generate_excel_buffer(df: pd.DataFrame) -> io.BytesIO:
    """Генерирует Excel файл в памяти."""
    # Очищаем DataFrame перед записью
    df = sanitize_dataframe(df.copy())

    buf = io.BytesIO()
    try:
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='assign')
            ws = writer.sheets['assign']
            for idx, col in enumerate(df.columns, start=1):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                ws.column_dimensions[get_column_letter(idx)].width = max_len
        buf.seek(0)
        logging.info("Excel буфер успешно сформирован")
        return buf
    except Exception as e:
        logging.error(f"Ошибка при формировании Excel-буфера: {e}")
        raise